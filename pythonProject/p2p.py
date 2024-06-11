import requests
import hashlib
import hmac
import time
from openpyxl import Workbook
import locale

def get_binance_c2c_order_history(api_key, secret_key):
    base_url = 'https://api.binance.com'
    endpoint = '/sapi/v1/c2c/orderMatch/listUserOrderHistory'
    # Calculate start timestamp (20 days ago) in milliseconds
    start_timestamp = int((time.time() - 2000 * 24 * 3600) * 1000)
    timestamp = int(time.time() * 1000)

    # Create the query string
    query_string = f"timestamp={timestamp}&startTime={start_timestamp}"

    # Create the signature
    signature = hmac.new(secret_key.encode(), query_string.encode(), hashlib.sha256).hexdigest()

    # Create request URL
    url = f"{base_url}{endpoint}?{query_string}&signature={signature}"

    # Set request headers
    headers = {
        'X-MBX-APIKEY': api_key
    }

    # Send GET request to Binance API
    response = requests.get(url, headers=headers)

    # Check if request was successful
    if response.status_code == 200:
        response_json = response.json()
        return response_json
    else:
        print(f"Error: {response.status_code}, {response.text}")
        return None



def save_to_excel(buy_orders, sell_orders, profit, filename):
    # Set the locale to format numbers with commas
    locale.setlocale(locale.LC_NUMERIC, 'en_US.UTF-8')

    # Create a new Excel workbook
    wb = Workbook()
    ws = wb.active

    # Write headers for buy orders
    headers = ["Order Status", "Total Price", "Trade Type", "Create Time", "Profit"]
    for col, header in enumerate(headers, start=1):
        ws.cell(row=1, column=col, value=header)

    # Write buy orders data starting from cell A2
    for i, order in enumerate(buy_orders, start=2):
        total_price = int(locale.atof(order["totalPrice"]))
        row = [
            order["orderStatus"],
            total_price,
            order["tradeType"],
            order["createTime"],
            ""  # Placeholder for profit, will be filled later
        ]
        for j, value in enumerate(row, start=1):
            ws.cell(row=i, column=j, value=value)

    # Write headers for sell orders starting from cell H1
    sell_headers = ["Order Status", "Total Price", "Trade Type", "Create Time", "Profit"]
    for col, header in enumerate(sell_headers, start=1):
        ws.cell(row=1, column=col+7, value=header)

    # Write sell orders data starting from cell H2
    for i, order in enumerate(sell_orders, start=2):
        total_price = int(locale.atof(order["totalPrice"]))
        row = [
            order["orderStatus"],
            total_price,
            order["tradeType"],
            order["createTime"],
            ""  # Placeholder for profit, will be filled later
        ]
        for j, value in enumerate(row, start=1):
            ws.cell(row=i, column=j+7, value=value)


    ws.cell(row=1, column=13, value="Profit")
    ws.cell(row=2, column=13, value=profit)


    wb.save(filename)


def count_profit(parsed_orders):
    profit = 0.0
    for order in parsed_orders:
        if order['tradeType'] == 'BUY':
            profit -= float(order['totalPrice'])
        else:
            profit += float(order['totalPrice'])
    return profit
def separate_buy_sell(parsed_orders):
    buy_orders = []
    sell_orders = []
    for order in parsed_orders:
        if order['tradeType'] == 'BUY':
            buy_orders.append(order)
        else:
            sell_orders.append(order)
    return buy_orders, sell_orders